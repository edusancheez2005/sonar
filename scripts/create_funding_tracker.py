"""
Generate Sonar Funding Opportunities Tracker Excel file
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════
# COLOR PALETTE
# ═══════════════════════════════════════════════
DARK_BG = "0A0E17"
CYAN = "00E5FF"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
HEADER_BG = "1B2838"
HEADER_FG = "00E5FF"
GREEN_BG = "E2EFDA"
GREEN_FG = "2E7D32"
YELLOW_BG = "FFF9C4"
YELLOW_FG = "F57F17"
RED_BG = "FFEBEE"
RED_FG = "C62828"
BLUE_BG = "E3F2FD"
BLUE_FG = "1565C0"
GRAY_BG = "F5F5F5"
PURPLE_BG = "F3E5F5"
PURPLE_FG = "6A1B9A"
ORANGE_BG = "FFF3E0"
ORANGE_FG = "E65100"

# Styles
header_font = Font(name="Calibri", bold=True, size=11, color=WHITE)
header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

title_font = Font(name="Calibri", bold=True, size=14, color="1565C0")
subtitle_font = Font(name="Calibri", bold=True, size=11, color="555555")

data_font = Font(name="Calibri", size=10)
data_align = Alignment(vertical="top", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Status fills
status_fills = {
    "SUBMITTED": PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid"),
    "DRAFT READY": PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid"),
    "NOT STARTED": PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type="solid"),
    "AWARDED": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
    "REJECTED": PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid"),
    "IN PROGRESS": PatternFill(start_color=BLUE_BG, end_color=BLUE_BG, fill_type="solid"),
    "SKIPPED": PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"),
}
status_fonts = {
    "SUBMITTED": Font(name="Calibri", bold=True, size=10, color=GREEN_FG),
    "DRAFT READY": Font(name="Calibri", bold=True, size=10, color=YELLOW_FG),
    "NOT STARTED": Font(name="Calibri", size=10, color="757575"),
    "AWARDED": Font(name="Calibri", bold=True, size=10, color="1B5E20"),
    "REJECTED": Font(name="Calibri", bold=True, size=10, color=RED_FG),
    "IN PROGRESS": Font(name="Calibri", bold=True, size=10, color=BLUE_FG),
    "SKIPPED": Font(name="Calibri", size=10, color="9E9E9E"),
}

priority_fills = {
    "Very High": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
    "High": PatternFill(start_color=BLUE_BG, end_color=BLUE_BG, fill_type="solid"),
    "Medium": PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid"),
    "Low": PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type="solid"),
}
priority_fonts = {
    "Very High": Font(name="Calibri", bold=True, size=10, color=GREEN_FG),
    "High": Font(name="Calibri", bold=True, size=10, color=BLUE_FG),
    "Medium": Font(name="Calibri", size=10, color=YELLOW_FG),
    "Low": Font(name="Calibri", size=10, color="757575"),
}


def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def style_data_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = thin_border


def apply_status_style(cell, status):
    for key in status_fills:
        if key in str(status).upper():
            cell.fill = status_fills[key]
            cell.font = status_fonts[key]
            return


def apply_priority_style(cell, priority):
    for key in priority_fills:
        if key.lower() in str(priority).lower():
            cell.fill = priority_fills[key]
            cell.font = priority_fonts[key]
            return


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════
# SHEET 1: APPLICATION TRACKER (main tracking sheet)
# ═══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Application Tracker"
ws1.sheet_properties.tabColor = "1565C0"

# Title
ws1.merge_cells("A1:J1")
ws1["A1"].value = "SONAR — Grant & Funding Application Tracker"
ws1["A1"].font = title_font
ws1["A1"].alignment = Alignment(vertical="center")
ws1.row_dimensions[1].height = 30

ws1.merge_cells("A2:J2")
ws1["A2"].value = f"Last updated: {datetime.now().strftime('%d %B %Y')} | Target: £50k–£500k | Stage: Pre-revenue"
ws1["A2"].font = subtitle_font
ws1.row_dimensions[2].height = 20

headers = [
    "Application", "Type", "Amount", "Equity/Cost", "Priority",
    "Status", "Date Submitted", "Deadline", "Response", "Notes"
]
for col, h in enumerate(headers, 1):
    ws1.cell(row=4, column=col, value=h)
style_header_row(ws1, 4, len(headers))
ws1.auto_filter.ref = f"A4:J4"
ws1.freeze_panes = "A5"

applications = [
    # [name, type, amount, equity, priority, status, submitted, deadline, response, notes]
    ["Solana Foundation Grant", "Crypto ecosystem grant", "$15,000", "Non-dilutive", "Very High",
     "SUBMITTED", "7 March 2026", "Rolling", "", "Draft: SOLANA_GRANT_DRAFT.md. Awaiting response (4 days)."],
    ["Superteam UK Microgrant", "Bounties / challenges", "$10,000 USDG", "Non-dilutive", "Low",
     "SKIPPED", "", "Rolling", "", "Not a traditional grant — it's a bounty/challenges board. Deprioritised."],
    ["Base Builder Grants", "Crypto ecosystem grant", "1–5 ETH", "Non-dilutive", "Very High",
     "DRAFT READY", "", "Rolling (ship first)", "", "Draft: BASE_BUILDER_GRANT_DRAFT.md. Also register for Builder Rewards."],
    ["Eagle Labs Funding Readiness", "Programme (non-cash)", "Non-cash", "None", "High",
     "NOT STARTED", "", "13 March 2026", "", "Quick win for tightening investor narrative."],
    ["HMRC SEIS/EIS Advance Assurance", "Tax scheme enabler", "Unlocks £250k SEIS", "None", "Very High",
     "NOT STARTED", "", "Rolling (target: 13 March)", "", "Must do before angel outreach. Validate crypto 'qualifying trade'."],
    ["Arbitrum Foundation Grants", "Crypto ecosystem grant", "Not stated", "Non-dilutive", "High",
     "NOT STARTED", "", "Rolling", "", "Infrastructure & Tools lane. Analytics tooling fits."],
    ["BNB Chain Grants", "Crypto ecosystem grant", "Up to $200k", "Non-dilutive", "High",
     "NOT STARTED", "", "Rolling", "", "Already track BSC. Package as open tooling + dashboards."],
    ["ARIA Scaling Trust", "Government grant", "£100k–£3m", "Non-dilutive", "Medium",
     "NOT STARTED", "", "24 March 2026", "", "Only if ORCA 2.0 fits 'trustworthy agent' framing. High effort."],
    ["Base Builder Rewards", "Recurring builder reward", "2 ETH/week", "Non-dilutive", "High",
     "NOT STARTED", "", "Rolling", "", "Register at builderscore.xyz. Maintain Builder Score."],
    ["Ethereum Foundation ESP", "Crypto ecosystem grant", "Scope-based", "Non-dilutive", "Medium",
     "NOT STARTED", "", "Rolling", "", "If you OSS whale classification components."],
    ["Barclays Eagle Labs / Capital Enterprise", "Funding readiness", "Non-cash", "None", "High",
     "NOT STARTED", "", "13 March 2026", "", "Cohort programme. Tighten investor narrative."],
    ["NatWest FinTech Programme 2026", "Accelerator", "Varies", "Varies", "Medium",
     "NOT STARTED", "", "Check deadline", "", "Position as AI intelligence for financial services."],
    ["Techstars London", "Accelerator", "$220k (5% + SAFE)", "Equity + SAFE", "Medium",
     "NOT STARTED", "", "April 2026 (verify)", "", "Solo founder not disqualifying. Need traction."],
    ["Entrepreneurs First Bridge (Summer 2026)", "Talent accelerator", "Up to $250k", "SAFE", "Low",
     "NOT STARTED", "", "6 April 2026 (early)", "", "⚠️ Full-time in-person. High commitment."],
    ["Innovate UK Innovation Loans R26", "Innovation loan", "£100k–£5m", "Debt", "Low",
     "NOT STARTED", "", "29 April 2026", "", "Only if repayment story is credible."],
    ["ARIA Rolling Opportunity Seeds", "Government grant", "Up to £500k", "Non-dilutive", "Low",
     "NOT STARTED", "", "Rolling", "", "Check monthly for new opportunity spaces."],
    ["Starknet Seed Grants", "Crypto ecosystem grant", "Up to $25k (STRK)", "Non-dilutive", "Medium",
     "NOT STARTED", "", "Rolling", "", "If adding Starknet coverage."],
    ["Osmosis Grants", "Crypto ecosystem grant", "Varies", "Non-dilutive", "Medium",
     "NOT STARTED", "", "Rolling", "", "Explicitly funds analytics tooling. If expanding to Cosmos."],
    ["KTP 2026-27 Round 1", "Collaboration grant", "Varies", "Cost-share", "Low",
     "NOT STARTED", "", "1 April 2026", "", "Led by academic institutions. Good for hiring KTP associate."],
    ["Hub71 Access (Cohort 20)", "UAE incentive programme", "In-kind + cash", "Equity", "Medium",
     "NOT STARTED", "", "2 August 2026", "", "Requires UAE operational readiness."],
    ["Start Up Loans", "Govt-backed loan", "Up to £25k", "Debt (6% fixed)", "Low",
     "NOT STARTED", "", "Rolling", "", "Bridge for early engineering/marketing spend."],
    # ── ACCELERATORS ──
    ["Y Combinator (YC)", "Accelerator", "$500k ($125k + $375k SAFE)", "7% equity + SAFE", "High",
     "NOT STARTED", "", "Batch-based (check YC.com)", "", "Top-tier. Remote-friendly. Crypto/fintech fits. Highly competitive but massive signal value."],
    ["a16z Crypto Startup School (CSS)", "Accelerator", "Equity-free (some cohorts) / varies", "Varies by programme", "High",
     "NOT STARTED", "", "Cohort-based (check a16zcrypto.com)", "", "Crypto-native accelerator. Network access to a16z portfolio. Verify current programme format."],
    ["Alliance DAO", "Crypto accelerator", "$250k–$1M SAFE", "Equity (SAFE)", "Very High",
     "NOT STARTED", "", "Cohort-based (check alliance.xyz)", "", "Best crypto-native accelerator. DeFi/trading/analytics fit. NYC-based, remote possible."],
    ["Seedcamp", "Accelerator (Europe)", "Up to $500k", "~7.5% equity", "High",
     "NOT STARTED", "", "Rolling", "", "Europe's top pre-seed fund. London-based. Strong fintech/data portfolio."],
    ["Antler", "Accelerator", "$100k–$250k", "~10% equity", "Medium",
     "NOT STARTED", "", "Cohort-based (check antler.co)", "", "London cohorts available. Good for solo founders looking for cofounders."],
    ["500 Global", "Accelerator", "$150k", "6% equity", "Medium",
     "NOT STARTED", "", "Batch-based (check 500.co)", "", "Global programme, remote-friendly. Broad network but less crypto-specific."],
    ["Colosseum (Solana)", "Crypto accelerator", "$250k", "Equity/SAFE", "High",
     "NOT STARTED", "", "Hackathon → accelerator pipeline", "", "Solana-native. Hackathon winners get accelerator access. Strong if Solana grant lands."],
    ["Outlier Ventures Base Camp", "Web3 accelerator", "Varies", "Equity; some tracks: 6% token supply", "Medium",
     "NOT STARTED", "", "Check availability", "", "⚠️ Many tracks assume token economics. Not aligned unless you build token components."],
    ["Founders Factory", "Accelerator (London)", "Up to £250k", "~5-8% equity", "Medium",
     "NOT STARTED", "", "Rolling (check foundersfactory.com)", "", "London-based. Corporate-backed programmes. Strong operational support."],
    ["Plug and Play FinTech", "Accelerator", "No investment (platform)", "None", "Medium",
     "NOT STARTED", "", "Batch-based", "", "No equity taken. Access to 500+ corporate partners. Strong fintech vertical."],
    ["Binance Labs (MVB)", "Crypto accelerator", "Up to $500k", "Equity/token", "Medium",
     "NOT STARTED", "", "Cohort-based (check labs.binance.com)", "", "Most Valuable Builder (MVB) programme. BNB Chain focus. Token expectations possible."],
    ["Blockchain Valley Ventures (BVV)", "Crypto accelerator", "CHF 50k–250k", "Equity", "Low",
     "NOT STARTED", "", "Rolling", "", "Swiss-based. Crypto/Web3 focus. Good for European expansion narrative."],
]

for i, app in enumerate(applications):
    row = 5 + i
    for col, val in enumerate(app, 1):
        ws1.cell(row=row, column=col, value=val)
    style_data_row(ws1, row, len(headers))
    # Apply status styling
    apply_status_style(ws1.cell(row=row, column=6), app[5])
    # Apply priority styling
    apply_priority_style(ws1.cell(row=row, column=5), app[4])

set_col_widths(ws1, [35, 22, 18, 15, 12, 15, 16, 22, 12, 55])

# ═══════════════════════════════════════════════════════
# SHEET 2: WEEK-BY-WEEK ACTION PLAN
# ═══════════════════════════════════════════════════════
ws2 = wb.create_sheet("Weekly Action Plan")
ws2.sheet_properties.tabColor = "2E7D32"

ws2.merge_cells("A1:F1")
ws2["A1"].value = "Week-by-Week Action Plan — March–April 2026"
ws2["A1"].font = title_font
ws2.row_dimensions[1].height = 30

plan_headers = ["Week", "Action", "Why", "Deadline", "Status", "Notes"]
for col, h in enumerate(plan_headers, 1):
    ws2.cell(row=3, column=col, value=h)
style_header_row(ws2, 3, len(plan_headers))
ws2.freeze_panes = "A4"

actions = [
    # Week of 2 March
    ["2 Mar", "Write one-page grant memo (open-source per chain)", "Supports ecosystem grants", "—", "DONE", ""],
    ["2 Mar", "Write investor one-pager (SEIS/EIS-ready)", "Needed for angel conversations", "—", "DONE", ""],
    ["2 Mar", "Decide on Innovate UK Loans R25", "Closes 4 March", "4 March 2026", "SKIPPED", "Pre-revenue, can't show repayment"],
    # Week of 9 March (CURRENT WEEK)
    ["9 Mar ★", "Submit Base Builder Grants (1-5 ETH) ← moved up", "Explicitly lists analytics dashboards; best crypto grant fit", "Rolling", "DRAFT READY", "BASE_BUILDER_GRANT_DRAFT.md"],
    ["9 Mar ★", "Submit Base Builder Grants (1-5 ETH)", "Explicitly lists analytics dashboards", "Rolling", "DRAFT READY", "BASE_BUILDER_GRANT_DRAFT.md"],
    ["9 Mar ★", "Apply Eagle Labs Funding Readiness", "Non-cash but high leverage", "13 March 2026", "NOT STARTED", "enterprisenation.com"],
    ["9 Mar ★", "Attend FinTech Investment Breakfast", "10 warm intros → 3 follow-ups", "11 March 2026", "NOT STARTED", "Barclays Innovation Hub"],
    ["9 Mar ★", "Submit HMRC SEIS/EIS advance assurance", "Unlocks angel conversations", "By 13 March", "NOT STARTED", "gov.uk"],
    ["9 Mar ★", "Competitor gap analysis", "Know where we win and lose", "—", "DONE", "COMPETITOR_GAP_ANALYSIS.md"],
    # Week of 16 March
    ["16 Mar", "Draft ARIA Scaling Trust proposal", "£100k–£3m non-dilutive", "24 March 2026", "NOT STARTED", "Only if ORCA fits 'trustworthy agent'"],
    ["16 Mar", "Submit Arbitrum Foundation grant", "Infra & Tools lane", "Rolling", "NOT STARTED", ""],
    ["16 Mar", "Submit BNB Chain grant", "Up to $200k; already track BSC", "Rolling", "NOT STARTED", ""],
    ["16 Mar", "Build CRM: 30 UK angels + 15 crypto funds", "Pipeline for April events", "—", "NOT STARTED", ""],
    # Week of 23 March
    ["23 Mar", "Submit ARIA Scaling Trust (if proceeding)", "Non-dilutive grant", "24 March 2026", "NOT STARTED", ""],
    ["23 Mar", "Book 10-15 investor meetings for April", "Stack around IFGS + TOKEN2049", "—", "NOT STARTED", ""],
    ["23 Mar", "Build 6-slide conference deck", "For all April events", "—", "NOT STARTED", ""],
    ["23 Mar", "Pre-write cold DM template", "Outreach efficiency", "—", "NOT STARTED", ""],
    # Week of 30 March
    ["30 Mar", "Decide on EF Bridge Residency", "High commitment decision", "6 April 2026", "NOT STARTED", ""],
    ["30 Mar", "Prepare Techstars London app", "Verify deadline", "April 2026", "NOT STARTED", ""],
    # Week of 6 April
    ["6 Apr", "Prepare Dubai trip", "Schedule meeting blocks", "—", "NOT STARTED", ""],
    ["6 Apr", "Create Dubai pitch version", "UAE presence + compliance", "—", "NOT STARTED", ""],
    # Week of 13 April
    ["13 Apr", "Attend AIM Investment Summit (Dubai)", "25 convos → 8 leads → 4 follow-ups", "13-15 April", "NOT STARTED", ""],
    # Week of 20 April
    ["20 Apr", "Attend IFGS (London)", "Best UK fintech investor concentration", "21 April", "NOT STARTED", ""],
    ["20 Apr", "Send follow-ups within 48h", "Compounding effect", "—", "NOT STARTED", ""],
    # Week of 27 April
    ["27 Apr", "Attend TOKEN2049 Dubai", "Highest crypto fund density", "29-30 April", "NOT STARTED", ""],
    ["27 Apr", "Submit Innovate UK R26 (if applying)", "Don't miss deadline", "29 April 2026", "NOT STARTED", ""],
    ["27 Apr", "Finalise chain grant submissions", "BNB + Base + Arbitrum closed", "—", "NOT STARTED", ""],
]

for i, action in enumerate(actions):
    row = 4 + i
    for col, val in enumerate(action, 1):
        ws2.cell(row=row, column=col, value=val)
    style_data_row(ws2, row, len(plan_headers))
    apply_status_style(ws2.cell(row=row, column=5), action[4])
    # Bold current week
    if "★" in str(action[0]):
        ws2.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=10, color=BLUE_FG)

set_col_widths(ws2, [12, 45, 35, 20, 15, 35])

# ═══════════════════════════════════════════════════════
# SHEET 3: CRYPTO ECOSYSTEM GRANTS
# ═══════════════════════════════════════════════════════
ws3 = wb.create_sheet("Crypto Grants")
ws3.sheet_properties.tabColor = "FF6F00"

ws3.merge_cells("A1:H1")
ws3["A1"].value = "Crypto Ecosystem Grants — Full List"
ws3["A1"].font = title_font
ws3.row_dimensions[1].height = 30

grant_headers = ["Program", "Ecosystem", "Amount", "Analytics Eligible?", "Requirements", "URL", "Priority", "Status"]
for col, h in enumerate(grant_headers, 1):
    ws3.cell(row=3, column=col, value=h)
style_header_row(ws3, 3, len(grant_headers))
ws3.auto_filter.ref = f"A3:H3"
ws3.freeze_panes = "A4"

grants = [
    ["Base Builder Grants", "Base (L2)", "1–5 ETH", "YES — explicitly", "Build on Base mainnet; document work", "docs.base.org/get-started/get-funded", "Very High", "DRAFT READY"],
    ["Base Builder Rewards", "Base (L2)", "2 ETH/week", "YES", "Maintain Builder Score", "builderscore.xyz", "Very High", "NOT STARTED"],
    ["Superteam UK Microgrants", "Solana", "Up to $10k USDG", "YES", "UK-only; bounty/challenge format", "superteam.fun/earn/grants/solana-foundation-uk-grants/", "Low", "SKIPPED — not a proper grant, just challenges"],
    ["Solana Foundation (standard/convertible)", "Solana", "$15,000", "Medium (public goods)", "Explain public good; OSS spirit", "solana.org/grants-funding", "Very High", "SUBMITTED"],
    ["BNB Chain Grants", "BNB Chain", "Up to $200k", "YES (tooling)", "Align to wishlist; two-stage review", "bnbchain.org/en/grants", "High", "NOT STARTED"],
    ["Arbitrum Foundation Grants", "Arbitrum (L2)", "Not stated", "YES (tools lane)", "Match initiative focus; milestones", "arbitrum.foundation/grants", "High", "NOT STARTED"],
    ["Arbitrum DAO Grant Program S3", "Arbitrum (L2)", "Varies", "YES", "Via Questbook portal", "arbitrum.questbook.app", "High", "NOT STARTED"],
    ["Ethereum Foundation ESP", "Ethereum", "Scope-based", "YES (if OSS tooling)", "Align to Wishlist/RFP; KYC", "esp.ethereum.foundation/applicants", "Medium-High", "NOT STARTED"],
    ["Starknet Seed Grants", "Starknet", "Up to $25k (STRK)", "YES", "MVP/proof-of-concept", "starknet.io/grants/seed-grants/", "Medium", "NOT STARTED"],
    ["Osmosis Grants", "Cosmos/Osmosis", "Varies", "YES — explicitly", "Contribute to Osmosis ecosystem", "grants.osmosis.zone", "Medium", "NOT STARTED"],
    ["Avalanche Builder Hub", "Avalanche", "Programme-specific", "YES (tooling)", "Programme-specific", "build.avax.network/grants", "Medium", "NOT STARTED"],
    ["NEAR Funding Hub", "NEAR", "Programme-specific", "YES (tooling)", "Build on NEAR", "near.org/funding", "Medium", "NOT STARTED"],
    ["Optimism Retro Funding", "Optimism (L2)", "Varies by round", "YES (after impact)", "Demonstrate impact first", "gov.optimism.io/c/grants/retrofunding/46", "Medium", "NOT STARTED"],
    ["Tezos Foundation Grants", "Tezos", "Not fixed", "YES", "Quarterly review cycle", "grants.tezos.foundation", "Medium", "NOT STARTED"],
    ["Interchain Foundation", "Cosmos", "Not fixed", "YES (tooling)", "Align to ICF objectives", "apply.interchain.io", "Medium", "NOT STARTED"],
    ["Project Catalyst (Cardano)", "Cardano", "Round-based", "YES", "Community vote process", "projectcatalyst.io", "Medium", "NOT STARTED"],
    ["Gitcoin Grants", "Multi-ecosystem", "Round-specific", "YES", "Public goods impact", "grants.gitcoin.co", "Medium", "NOT STARTED"],
    ["Starknet Growth Grants", "Starknet", "Up to $1m", "YES", "Later-stage; strong fit", "starknet.io/grants/growth-grants/", "Low-Medium", "NOT STARTED"],
]

for i, grant in enumerate(grants):
    row = 4 + i
    for col, val in enumerate(grant, 1):
        ws3.cell(row=row, column=col, value=val)
    style_data_row(ws3, row, len(grant_headers))
    apply_priority_style(ws3.cell(row=row, column=7), grant[6])
    apply_status_style(ws3.cell(row=row, column=8), grant[7])

set_col_widths(ws3, [35, 16, 18, 18, 35, 45, 14, 15])

# ═══════════════════════════════════════════════════════
# SHEET 4: UK OPPORTUNITIES
# ═══════════════════════════════════════════════════════
ws4 = wb.create_sheet("UK Opportunities")
ws4.sheet_properties.tabColor = "6A1B9A"

ws4.merge_cells("A1:H1")
ws4["A1"].value = "UK Funding Opportunities"
ws4["A1"].font = title_font
ws4.row_dimensions[1].height = 30

uk_headers = ["Opportunity", "Type", "Amount/Benefit", "Equity/Cost", "Deadline", "URL", "Priority", "Status"]
for col, h in enumerate(uk_headers, 1):
    ws4.cell(row=3, column=col, value=h)
style_header_row(ws4, 3, len(uk_headers))
ws4.auto_filter.ref = f"A3:H3"
ws4.freeze_panes = "A4"

uk_opps = [
    ["HMRC SEIS/EIS Advance Assurance", "Tax enabler", "Investor confidence", "None", "Rolling (target 13 Mar)", "gov.uk SEIS advance assurance", "Very High", "NOT STARTED"],
    ["SEIS (£250k cap)", "Tax relief for investors", "Up to £250,000", "Equity", "Rolling", "gov.uk SEIS", "Very High", "NOT STARTED"],
    ["EIS (second close)", "Tax relief for investors", "Up to £5m/year", "Equity", "Rolling", "gov.uk EIS", "High", "NOT STARTED"],
    ["Eagle Labs Funding Readiness", "Programme (non-cash)", "Cohort programme", "None", "13 March 2026", "enterprisenation.com", "High", "NOT STARTED"],
    ["ARIA Scaling Trust", "Government grant", "£100k–£3m", "Non-dilutive", "24 March 2026", "aria.org.uk scaling trust", "Medium", "NOT STARTED"],
    ["ARIA Rolling Seeds", "Government grant", "Up to £500k", "Non-dilutive", "Rolling", "aria.org.uk/funding-opportunities", "Low", "NOT STARTED"],
    ["NatWest FinTech Programme", "Accelerator", "Varies", "Varies", "Check deadline", "natwestgroup.com fintech 2026", "Medium", "NOT STARTED"],
    ["FinTech Innovation Lab London", "Accelerator", "Mentor programme", "Usually no equity", "Check closing date", "fintechinnovationlab.com/london", "Medium", "NOT STARTED"],
    ["Techstars London", "Accelerator", "$220k (5% + SAFE)", "Equity + SAFE", "April 2026 (verify)", "techstars.com/accelerators/london", "Medium", "NOT STARTED"],
    ["Entrepreneurs First Bridge", "Talent accelerator", "Up to $250k", "SAFE", "6 April 2026 (early)", "apply.joinef.com", "Low", "NOT STARTED"],
    ["Innovate UK Loans R26", "Innovation loan", "£100k–£5m", "Debt", "29 April 2026", "apply-for-innovation-funding.service.gov.uk", "Low", "NOT STARTED"],
    ["Innovate UK Business Growth", "Support (non-cash)", "1:1 support", "None", "Rolling", "iuk-business-connect.org.uk", "Medium", "NOT STARTED"],
    ["KTP 2026-27 Round 1", "Collaboration grant", "Varies", "Cost-share", "1 April 2026", "apply-for-innovation-funding.service.gov.uk", "Low", "NOT STARTED"],
    ["Start Up Loans", "Govt loan", "Up to £25k", "Debt (6%)", "Rolling", "gov.uk/apply-start-up-loan", "Low", "NOT STARTED"],
    ["Level39 (Canary Wharf)", "Incubator/community", "Office + events", "Membership", "Rolling", "level39.co", "Low", "NOT STARTED"],
]

for i, opp in enumerate(uk_opps):
    row = 4 + i
    for col, val in enumerate(opp, 1):
        ws4.cell(row=row, column=col, value=val)
    style_data_row(ws4, row, len(uk_headers))
    apply_priority_style(ws4.cell(row=row, column=7), opp[6])
    apply_status_style(ws4.cell(row=row, column=8), opp[7])

set_col_widths(ws4, [35, 20, 20, 15, 22, 40, 14, 15])

# ═══════════════════════════════════════════════════════
# SHEET 5: UAE OPPORTUNITIES
# ═══════════════════════════════════════════════════════
ws5 = wb.create_sheet("UAE Opportunities")
ws5.sheet_properties.tabColor = "E65100"

ws5.merge_cells("A1:H1")
ws5["A1"].value = "UAE / Dubai Funding Opportunities"
ws5["A1"].font = title_font
ws5.row_dimensions[1].height = 30

uae_headers = ["Opportunity", "Type", "Amount/Benefit", "Equity/Cost", "Deadline", "URL", "Priority", "Status"]
for col, h in enumerate(uae_headers, 1):
    ws5.cell(row=3, column=col, value=h)
style_header_row(ws5, 3, len(uae_headers))
ws5.auto_filter.ref = f"A3:H3"
ws5.freeze_panes = "A4"

uae_opps = [
    ["Hub71 Access (Cohort 20)", "Incentive programme", "In-kind + cash", "Equity", "2 August 2026", "hub71.com/program/access-programme", "Medium", "NOT STARTED"],
    ["MBRIF Guarantee Scheme", "Govt-backed financing", "Credit guarantees", "No equity; debt", "Rolling", "mbrif.ae/guarantee-scheme-2", "Medium", "NOT STARTED"],
    ["in5 Tech", "Incubator", "Workspace + exposure", "Membership costs", "Rolling", "infive.ae/in5-tech", "Low", "NOT STARTED"],
    ["Dubai Future Accelerators", "Public-private accelerator", "Pilots + partnerships", "Usually non-dilutive", "Cohort-based", "dubaifuture.ae/accelerators", "Low", "NOT STARTED"],
    ["DIFC Innovation Licence", "Free zone licence", "USD 1,500/year", "Licence cost", "Live offer", "landing.difc.ae/innovation-license-offer", "Medium", "NOT STARTED"],
    ["DFSA Innovation Testing Licence", "Regulatory sandbox", "Restricted testing", "Regulatory fees", "Open window", "services.dfsa.ae ITL", "Low", "NOT STARTED"],
    ["DMCC Crypto Centre", "Free zone + ecosystem", "Crypto licence ~AED 31K", "Licence cost", "Rolling", "dmcc.ae/ecosystems/crypto-and-blockchain", "Low", "NOT STARTED"],
    ["ADGM Digital Assets Framework", "Free zone + regulatory", "Digital assets framework", "Licence cost", "Rolling", "adgm.com/digital-assets", "Low", "NOT STARTED"],
    # MENA VCs
    ["Cypher Capital", "Crypto VC", "Seed fund", "Equity", "Rolling", "cyphercapital.com", "Medium", "NOT STARTED"],
    ["Shorooq Partners", "VC", "Seed to growth", "Equity", "Rolling", "shorooq.com", "Medium", "NOT STARTED"],
    ["VentureSouq", "VC", "FinTech focus", "Equity", "Rolling", "venturesouq.com", "Medium", "NOT STARTED"],
    ["MEVP", "VC", "$200k–$1m tickets", "Equity", "Rolling", "mevp.com/funds", "Medium", "NOT STARTED"],
    ["BECO Capital", "VC", "Inception to growth", "Equity", "Rolling", "becocapital.com", "Low", "NOT STARTED"],
    ["Wamda Capital", "VC", "Seed/early-stage", "Equity", "Rolling", "wamdacapital.com", "Low", "NOT STARTED"],
    ["Global Ventures", "VC", "Emerging markets", "Equity", "Rolling", "global.vc", "Low", "NOT STARTED"],
]

for i, opp in enumerate(uae_opps):
    row = 4 + i
    for col, val in enumerate(opp, 1):
        ws5.cell(row=row, column=col, value=val)
    style_data_row(ws5, row, len(uae_headers))
    apply_priority_style(ws5.cell(row=row, column=7), opp[6])
    apply_status_style(ws5.cell(row=row, column=8), opp[7])

set_col_widths(ws5, [35, 22, 22, 15, 18, 40, 14, 15])

# ═══════════════════════════════════════════════════════
# SHEET 6: EVENTS & CONFERENCES
# ═══════════════════════════════════════════════════════
ws6 = wb.create_sheet("Events")
ws6.sheet_properties.tabColor = "C62828"

ws6.merge_cells("A1:H1")
ws6["A1"].value = "Pitch Events & Conferences — 2026"
ws6["A1"].font = title_font
ws6.row_dimensions[1].height = 30

event_headers = ["Event", "Date", "Location", "Cost", "What You Get", "Priority", "Attending?", "Notes"]
for col, h in enumerate(event_headers, 1):
    ws6.cell(row=3, column=col, value=h)
style_header_row(ws6, 3, len(event_headers))
ws6.freeze_panes = "A4"

events = [
    # UK
    ["FinTech Investment Breakfast", "11 March 2026", "Barclays Innovation Hub, London", "Check", "Targeted fintech investor networking", "High", "", "10 warm intros → 3 follow-ups"],
    ["IFGS 2026 (UK FinTech Week)", "21 April 2026", "Guildhall, London", "Paid", "Best UK fintech investor concentration", "Very High", "", "Top priority UK event"],
    ["London Tech Week 2026", "8–12 June 2026", "Olympia London", "Mixed", "Broad tech + investor ecosystem", "High", "", "Schedule side meetings"],
    ["FinTech Connect 2026", "1–2 December 2026", "ExCeL London", "Paid", "Fintech stages incl. Tokenize:LDN", "High", "", ""],
    # Dubai
    ["AIM Investment Summit", "13–15 April 2026", "Dubai World Trade Center", "Paid", "Broad MENA investor exposure", "Medium", "", "25 convos → 8 leads → 4 follow-ups"],
    ["TOKEN2049 Dubai", "29–30 April 2026", "Dubai", "Paid (premium)", "Highest density crypto funds/operators", "Very High", "", "Core ICP: crypto funds, exchanges"],
    ["Dubai FinTech Summit", "11–12 May 2026", "Madinat Jumeirah, Dubai", "Paid", "DIFC-led fintech + investor ecosystem", "High", "", "AI + market intelligence framing"],
    ["Expand North Star 2026", "8–10 December 2026", "Dubai (alongside GITEX)", "Paid", "Startup/investor connector", "High", "", "Late-year MENA pipeline"],
    ["GITEX GLOBAL 2026", "7–11 December 2026", "Dubai Exhibition Centre", "Paid", "Massive tech/investor exposure", "Medium-High", "", "Leverage for meetings, not footfall"],
]

for i, event in enumerate(events):
    row = 4 + i
    for col, val in enumerate(event, 1):
        ws6.cell(row=row, column=col, value=val)
    style_data_row(ws6, row, len(event_headers))
    apply_priority_style(ws6.cell(row=row, column=6), event[5])
    # Add separator between UK and Dubai
    if i == 3:
        for col in range(1, len(event_headers) + 1):
            ws6.cell(row=row, column=col).border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="medium", color="1565C0"),
            )

set_col_widths(ws6, [30, 18, 28, 12, 40, 14, 12, 35])

# ═══════════════════════════════════════════════════════
# SHEET 7: ACCELERATORS
# ═══════════════════════════════════════════════════════
ws7 = wb.create_sheet("Accelerators")
ws7.sheet_properties.tabColor = "00897B"

ws7.merge_cells("A1:K1")
ws7["A1"].value = "Accelerator & Incubator Programmes"
ws7["A1"].font = title_font
ws7.row_dimensions[1].height = 30

ws7.merge_cells("A2:K2")
ws7["A2"].value = "Global accelerators relevant for a pre-revenue crypto/fintech analytics startup (UK-based solo founder)"
ws7["A2"].font = subtitle_font
ws7.row_dimensions[2].height = 20

acc_headers = ["Programme", "Type", "Investment", "Equity/Cost", "Location", "Duration", "Deadline / Timing", "URL", "Crypto Fit?", "Priority", "Notes"]
for col, h in enumerate(acc_headers, 1):
    ws7.cell(row=4, column=col, value=h)
style_header_row(ws7, 4, len(acc_headers))
ws7.auto_filter.ref = f"A4:K4"
ws7.freeze_panes = "A5"

accelerators = [
    # [name, type, investment, equity, location, duration, deadline, url, crypto_fit, priority, notes]
    ["Y Combinator (YC)", "Top-tier accelerator", "$500k ($125k + $375k SAFE)", "7% equity + SAFE", "San Francisco (remote OK)", "3 months",
     "Batch-based (S26 apps ~Apr, W27 ~Oct)", "ycombinator.com/apply", "Yes (crypto batch exists)",
     "High", "Most prestigious. Massive alumni network (Coinbase, Stripe). Highly competitive (~1-2% acceptance). Solo founder OK. Apply with traction."],
    ["a16z Crypto Startup School (CSS)", "Crypto accelerator", "Varies by programme", "Varies (some equity-free)", "San Francisco / Remote", "12 weeks",
     "Cohort-based (check a16zcrypto.com)", "a16zcrypto.com", "Yes — crypto-native",
     "High", "Run by Andreessen Horowitz crypto team. Deep crypto expertise. Access to a16z portfolio + LPs. Verify current cohort format — has changed between years."],
    ["Alliance DAO", "Crypto accelerator", "$250k–$1M (SAFE)", "Equity (SAFE terms)", "New York (remote possible)", "12 weeks",
     "Cohort-based (check alliance.xyz)", "alliance.xyz", "Yes — best crypto fit",
     "Very High", "#1 crypto-native accelerator. Founded by ex-ConsenSys. DeFi/trading/analytics/data companies in portfolio. VERY strong for Sonar's profile. Alumni: Phantom, Tensor, etc."],
    ["Seedcamp", "Pre-seed VC + accelerator", "Up to $500k (often ~€100-200k first cheque)", "~7.5% equity", "London (Europe-wide)", "Ongoing support",
     "Rolling (apply anytime)", "seedcamp.com/apply", "Yes (fintech/data focus)",
     "High", "Europe's leading pre-seed fund. Based in London — easy for you. Strong fintech portfolio (TransferWise/Wise, Revolut early). Rolling applications = no batch timing pressure."],
    ["Antler", "Talent-first accelerator", "$100k–$250k", "~10% equity", "London + 25 global locations", "6 months",
     "London cohorts: multiple per year", "antler.co/apply", "Medium-High",
     "Medium", "Good for solo founders — they help find cofounders. London programme available. Less crypto-specific but fintech-friendly. Residency model (in-person)."],
    ["500 Global", "Global accelerator", "$150k", "6% equity", "San Francisco / Remote", "4 months",
     "Batch-based (multiple per year)", "500.co", "Medium",
     "Medium", "Large global portfolio (2,800+ companies). Less crypto-focused but broad network. Remote-friendly. Good for general startup support."],
    ["Techstars London", "Top-tier accelerator", "$220k (5% CEA + SAFE)", "5% equity + uncapped MFN SAFE", "London", "3 months",
     "London batch 2026 (verify app close date)", "techstars.com/accelerators/london", "Medium (fintech focus)",
     "Medium", "Already in funding tracker. Strong London network. Solo founder not disqualifying. Demo day investor access. Less crypto-specific."],
    ["Entrepreneurs First (EF)", "Talent-first accelerator", "Up to $250k (SAFE)", "SAFE investment", "London (in-person required)", "6 months",
     "Bridge Residency Summer 2026: early deadline 6 Apr 2026", "apply.joinef.com", "Medium",
     "Low", "Already in funding tracker. ⚠️ Full-time in-person required. High commitment. May conflict with running a live product."],
    ["Colosseum (Solana)", "Crypto accelerator", "$250k", "Equity/SAFE", "Remote", "~3 months post-hackathon",
     "Hackathon → accelerator pipeline", "colosseum.org", "Yes — Solana-native",
     "High", "Best path if Solana grant lands. Hackathon winners → accelerator. Online hackathons (Radar, Renaissance). Accelerator for top teams."],
    ["Outlier Ventures Base Camp", "Web3 accelerator", "Varies by track", "Equity; some: 6% future token supply", "London / Remote", "12 weeks",
     "Check current Base Camp tracks", "outlierventures.io/apply", "Yes — Web3 native",
     "Medium", "⚠️ Many tracks assume token economics (6% token supply). NOT aligned unless you intentionally build token/protocol components. Ask about non-token tracks."],
    ["Founders Factory", "Corporate-backed accelerator", "Up to £250k", "~5-8% equity", "London", "6 months",
     "Rolling (check foundersfactory.com)", "foundersfactory.com", "Medium (fintech vertical)",
     "Medium", "London-based. Backed by corporates (Aviva, L'Oréal, etc.). Strong operational support. Has a fintech vertical."],
    ["Plug and Play FinTech", "Corporate accelerator", "No direct investment", "None (platform model)", "Global (Silicon Valley HQ)", "3 months",
     "Batch-based (multiple verticals/year)", "plugandplaytechcenter.com/fintech", "Medium",
     "Medium", "No equity taken — purely platform/corporate access. 500+ corporate partners. Strong for B2B fintech partnerships. Less relevant for B2C crypto."],
    ["Binance Labs (MVB)", "Crypto accelerator", "Up to $500k", "Equity / token terms", "Remote / Global", "~3 months",
     "Cohort-based (check labs.binance.com)", "labs.binance.com", "Yes — BNB Chain native",
     "Medium", "Most Valuable Builder (MVB) programme. BNB Chain focus aligns with BNB grant. May expect token economics. Massive distribution via Binance ecosystem."],
    ["Blockchain Valley Ventures (BVV)", "Swiss crypto accelerator", "CHF 50k–250k", "Equity", "Zug, Switzerland", "Varies",
     "Rolling", "bvventures.com", "Yes — crypto",
     "Low", "Based in Crypto Valley (Zug). European crypto VC/accelerator. Smaller programme. Good for EU expansion narrative."],
    ["NatWest FinTech Programme", "Banking accelerator", "Varies", "Varies / not stated", "London", "Varies",
     "Open (check deadline)", "natwestgroup.com fintech 2026", "Medium (fintech, not crypto)",
     "Medium", "AI-focused fintech programme. Strong if positioned as 'AI intelligence for financial services'. Verify crypto comfort."],
    ["FinTech Innovation Lab London", "Enterprise fintech accelerator", "Mentor-driven (no direct investment)", "Usually no equity", "London", "12 weeks",
     "Check closing date", "fintechinnovationlab.com/london", "Medium (B2B fintech)",
     "Medium", "Strong for B2B fintech traction. Run by Accenture + leading banks. More enterprise-focused than B2C."],
    ["Cointelegraph Accelerator", "Crypto media + accelerator", "$50k–$250k", "Equity", "Remote", "12 weeks",
     "Rolling / cohort-based", "cointelegraph.com/accelerator", "Yes — crypto media",
     "Medium", "Crypto media exposure + capital. Good for marketing leverage. Smaller programme."],
    ["Encode Club Accelerator", "Web3 education + accelerator", "Up to $200k", "Equity", "Remote", "8 weeks",
     "Cohort-based (check encode.club)", "encode.club", "Yes — Web3",
     "Medium", "Education-first Web3 accelerator. Multi-chain. Good for developer community building."],
]

for i, acc in enumerate(accelerators):
    row = 5 + i
    for col, val in enumerate(acc, 1):
        ws7.cell(row=row, column=col, value=val)
    style_data_row(ws7, row, len(acc_headers))
    apply_priority_style(ws7.cell(row=row, column=10), acc[9])

set_col_widths(ws7, [35, 25, 28, 22, 25, 12, 32, 38, 18, 14, 60])

# ═══════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════
output_path = r"c:\Users\t-eduardos\OneDrive - Microsoft\Desktop\UI sonar\Sonar_Funding_Tracker.xlsx"
wb.save(output_path)
print(f"✅ Excel file saved to: {output_path}")
print(f"   Sheets: {wb.sheetnames}")
print(f"   Applications tracked: {len(applications)}")
print(f"   Crypto grants listed: {len(grants)}")
print(f"   Accelerators listed: {len(accelerators)}")
print(f"   Weekly actions: {len(actions)}")
